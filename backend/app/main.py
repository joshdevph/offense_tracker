import base64
import binascii
import hashlib
import hmac
import json
import os
import random
import secrets
import sqlite3
import time
from contextlib import contextmanager
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Literal

from fastapi import Depends, FastAPI, File, Header, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from pydantic import BaseModel, Field


APP_NAME = "Offense Tracker"
PORTAL_TOKEN_TTL_SECONDS = 8 * 60 * 60
PASSWORD_ITERATIONS = 120_000
STUDENT_COLUMNS = (
    "id, student_no, first_name, last_name, grade, section, adviser, status, created_at, updated_at"
)
DB_PATH = Path(os.getenv("DATABASE_PATH", "/app/data/offense_tracker.db"))
if os.getenv("DATABASE_PATH") is None:
    DB_PATH = Path(__file__).resolve().parents[2] / "data" / "offense_tracker.db"

app = FastAPI(title=APP_NAME, version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=os.getenv("CORS_ORIGINS", "*").split(","),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class StudentIn(BaseModel):
    student_no: str = Field(..., min_length=1)
    first_name: str = Field(..., min_length=1)
    last_name: str = Field(..., min_length=1)
    grade: str = Field(..., min_length=1)
    section: str = Field(..., min_length=1)
    adviser: str = ""
    status: Literal["Active", "Inactive"] = "Active"


class BulkDeleteIn(BaseModel):
    ids: list[int] = Field(..., min_length=1, max_length=500)


class StudentLoginIn(BaseModel):
    username: str = Field(..., min_length=1)
    password: str = Field(..., min_length=1)


class OffenseIn(BaseModel):
    code: str = Field(..., min_length=1)
    name: str = Field(..., min_length=1)
    category: Literal["Minor", "Major"]
    severity_points: int = Field(..., ge=1, le=10)
    recommended_action: str = ""
    active: bool = True


class IncidentIn(BaseModel):
    incident_date: date | None = None
    student_id: int
    offense_id: int
    action_taken: str = ""
    status: Literal["Open", "Parent Notified", "Under Review", "Resolved"] = "Open"
    reported_by: str = ""
    notes: str = ""


def dict_factory(cursor: sqlite3.Cursor, row: sqlite3.Row) -> dict:
    return {col[0]: row[idx] for idx, col in enumerate(cursor.description)}


@contextmanager
def db():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = dict_factory
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def one(conn: sqlite3.Connection, query: str, params: tuple = ()) -> dict | None:
    return conn.execute(query, params).fetchone()


def all_rows(conn: sqlite3.Connection, query: str, params: tuple = ()) -> list[dict]:
    return conn.execute(query, params).fetchall()


def default_student_password(student_no: str, last_name: str) -> str:
    return f"{student_no}{''.join(last_name.split())}"


def hash_password(password: str) -> str:
    salt = secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, PASSWORD_ITERATIONS)
    return f"{base64.urlsafe_b64encode(salt).decode()}.{base64.urlsafe_b64encode(digest).decode()}"


def verify_password(password: str, stored_hash: str) -> bool:
    try:
        encoded_salt, encoded_digest = stored_hash.split(".", 1)
        salt = base64.urlsafe_b64decode(encoded_salt)
        expected = base64.urlsafe_b64decode(encoded_digest)
    except (ValueError, TypeError):
        return False
    actual = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, PASSWORD_ITERATIONS)
    return hmac.compare_digest(actual, expected)


def token_secret() -> bytes:
    return os.getenv("PORTAL_SECRET_KEY", "offense-tracker-local-development-key").encode()


def encode_portal_token(student_id: int) -> str:
    payload = {
        "student_id": student_id,
        "exp": int(time.time()) + PORTAL_TOKEN_TTL_SECONDS,
    }
    encoded = base64.urlsafe_b64encode(
        json.dumps(payload, separators=(",", ":")).encode()
    ).decode().rstrip("=")
    signature = hmac.new(token_secret(), encoded.encode(), hashlib.sha256).digest()
    encoded_signature = base64.urlsafe_b64encode(signature).decode().rstrip("=")
    return f"{encoded}.{encoded_signature}"


def decode_portal_token(token: str) -> int:
    try:
        encoded, encoded_signature = token.split(".", 1)
        padding = "=" * (-len(encoded_signature) % 4)
        supplied_signature = base64.urlsafe_b64decode(encoded_signature + padding)
        expected_signature = hmac.new(token_secret(), encoded.encode(), hashlib.sha256).digest()
        if not hmac.compare_digest(supplied_signature, expected_signature):
            raise ValueError("Invalid signature")
        payload_padding = "=" * (-len(encoded) % 4)
        payload = json.loads(base64.urlsafe_b64decode(encoded + payload_padding))
        if int(payload["exp"]) < int(time.time()):
            raise ValueError("Expired token")
        return int(payload["student_id"])
    except (binascii.Error, KeyError, TypeError, ValueError, json.JSONDecodeError):
        raise HTTPException(status_code=401, detail="Your student portal session is invalid or has expired")


def portal_student_id(authorization: str = Header(default="")) -> int:
    scheme, _, token = authorization.partition(" ")
    if scheme.lower() != "bearer" or not token:
        raise HTTPException(status_code=401, detail="Student portal login is required")
    return decode_portal_token(token)


def ensure_student_passwords(conn: sqlite3.Connection) -> None:
    students = all_rows(
        conn,
        "SELECT id, student_no, last_name FROM students WHERE password_hash = ''",
    )
    for student in students:
        password = default_student_password(student["student_no"], student["last_name"])
        conn.execute(
            "UPDATE students SET password_hash = ? WHERE id = ?",
            (hash_password(password), student["id"]),
        )


MAX_IMPORT_BYTES = 5 * 1024 * 1024
IMPORT_HEADERS = {
    "student_number": "student_no",
    "student_id": "student_no",
    "lrn": "student_no",
    "firstname": "first_name",
    "lastname": "last_name",
    "grade_level": "grade",
    "class_section": "section",
    "advisor": "adviser",
    "date": "incident_date",
    "offense": "offense_code",
    "code": "offense_code",
    "action": "action_taken",
    "reporter": "reported_by",
}

MINOR_SANCTION_GUIDE = (
    "1st offense: Class Adviser oral reprimand/counselling; "
    "2nd offense: refer to School Discipline Officer for oral and written reprimand; "
    "3rd offense: refer to Guidance Counselor for school/home therapy; "
    "4th offense: three-day suspension with counselling; "
    "5th offense: refer to the Principal."
)
MAJOR_SANCTION_GUIDE = (
    "Refer to the School Discipline Officer, Guidance Counselor, and Principal "
    "for investigation and appropriate disciplinary action."
)
OFFICIAL_OFFENSES = [
    ("MAJ-001", "Bullying (Physical, Emotional, Mental, Cyber Bullying)", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MAJ-002", "Cheating / Dishonesty", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MAJ-003", "Stealing / Theft / Robbery", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MAJ-004", "Assaulting a teacher, fellow learner, or school authority", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MAJ-005", "Smoking inside the school campus or within 100 meters of the school perimeter", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MAJ-006", "Vandalism or destruction of school property", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MAJ-007", "Gambling of any sort", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MAJ-008", "Drinking intoxicants/liquor, entering under the influence, or bringing alcoholic beverages", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MAJ-009", "Carrying or concealing deadly weapons or instruments", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MAJ-010", "Extortion or asking money or in-kind items from others", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MAJ-011", "Fighting or causing injury", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MAJ-012", "Hazing whether outside or inside the school campus", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MAJ-013", "Sexual abuse, immorality, or illicit relationships", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MAJ-014", "Instigating, leading, or participating in activities leading to stoppage of classes", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MAJ-015", "Preventing or threatening learners, faculty, or school authorities from discharging duties", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MAJ-016", "Forging or tampering with school records or transfer forms", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MAJ-017", "Truancy (Jumping over the fence)", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MAJ-018", "Joining fraternities, sororities, or gangs", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MAJ-019", "Watching or keeping pornographic materials", "Major", 5, MAJOR_SANCTION_GUIDE),
    ("MIN-001", "Absenteeism, cutting of classes, and tardiness", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-002", "Failure to wear prescribed uniform", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-003", "Wearing caps inside the school campus, especially inside the classroom", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-004", "Wearing civilian clothes on uniform days", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-005", "Wearing earrings for boys or multiple earrings for girls", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-006", "Outlandish attires, over-accessories, or body piercing", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-007", "Use of profane language or swearing", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-008", "Littering inside the school campus", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-009", "Using cellphones and other gadgets during class hours", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-010", "PDA (Public Display of Affection)", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-011", "Unruly behavior during assemblies, school activities, religious services, or flag ceremony", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-012", "Going to restricted places", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-013", "Refusal to wear or display school ID", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-014", "Placing stickers and other objects on school ID", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-015", "Wearing another student's ID", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-016", "Spending for personal use funds entrusted to him or her", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-017", "Refusing to obey student leaders discharging duty or representing authority", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-018", "Destroying another learner's belongings or blocking another person's path", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-019", "Disrespect to the Philippine Flag or singing of the National Anthem", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-020", "Spitting everywhere", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-021", "Putting on make-up and lipstick", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-022", "Urinating elsewhere or in inappropriate places", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-023", "Loitering or staying inside/outside the school campus during class hours", "Minor", 1, MINOR_SANCTION_GUIDE),
    ("MIN-024", "Other analogous acts that may endanger or threaten learners or school personnel", "Minor", 1, MINOR_SANCTION_GUIDE),
]
LEGACY_OFFENSE_CODE_MAP = {
    "MIN-003": "MIN-011",
    "MIN-004": "MIN-024",
    "MIN-005": "MIN-009",
    "MIN-006": "MIN-023",
    "MAJ-002": "MAJ-011",
    "MAJ-003": "MAJ-006",
    "MAJ-004": "MAJ-002",
    "MAJ-005": "MAJ-017",
    "MAJ-006": "MAJ-009",
}


def clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_header(value: object) -> str:
    header = clean_text(value).lower()
    normalized = "".join(char if char.isalnum() else "_" for char in header)
    normalized = "_".join(part for part in normalized.split("_") if part)
    return IMPORT_HEADERS.get(normalized, normalized)


def parse_workbook(contents: bytes) -> tuple[list[str], list[tuple[int, dict]]]:
    if not contents:
        raise HTTPException(status_code=400, detail="The selected Excel file is empty")
    if len(contents) > MAX_IMPORT_BYTES:
        raise HTTPException(status_code=413, detail="Excel files must be 5 MB or smaller")
    try:
        workbook = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Unable to read this file. Please upload a valid .xlsx workbook") from exc

    try:
        worksheet = workbook.active
        values = worksheet.iter_rows(values_only=True)
        first_row = next(values, None)
        if not first_row:
            raise HTTPException(status_code=400, detail="The first worksheet is empty")
        headers = [normalize_header(value) for value in first_row]
        records: list[tuple[int, dict]] = []
        for row_number, row in enumerate(values, start=2):
            record = {
                header: value
                for header, value in zip(headers, row)
                if header
            }
            if any(clean_text(value) for value in record.values()):
                records.append((row_number, record))
        return headers, records
    finally:
        workbook.close()


def require_import_headers(headers: list[str], required: set[str]) -> None:
    missing = sorted(required.difference(headers))
    if missing:
        readable = ", ".join(item.replace("_", " ").title() for item in missing)
        raise HTTPException(status_code=400, detail=f"Missing required Excel columns: {readable}")


def import_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        parsed = from_excel(value)
        return parsed.date() if isinstance(parsed, datetime) else parsed

    text = clean_text(value)
    if not text:
        raise ValueError("Incident Date is required")
    try:
        return date.fromisoformat(text)
    except ValueError:
        pass
    for pattern in ("%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise ValueError("Incident Date must use YYYY-MM-DD or a recognized Excel date")


def init_db() -> None:
    with db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS students (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_no TEXT NOT NULL UNIQUE,
                first_name TEXT NOT NULL,
                last_name TEXT NOT NULL,
                grade TEXT NOT NULL,
                section TEXT NOT NULL,
                adviser TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'Active',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS offenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                category TEXT NOT NULL CHECK(category IN ('Minor', 'Major')),
                severity_points INTEGER NOT NULL,
                recommended_action TEXT NOT NULL DEFAULT '',
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS incidents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                incident_no TEXT NOT NULL UNIQUE,
                incident_date TEXT NOT NULL,
                student_id INTEGER NOT NULL,
                offense_id INTEGER NOT NULL,
                action_taken TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'Open',
                reported_by TEXT NOT NULL DEFAULT '',
                notes TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(student_id) REFERENCES students(id) ON DELETE RESTRICT,
                FOREIGN KEY(offense_id) REFERENCES offenses(id) ON DELETE RESTRICT
            );

            CREATE INDEX IF NOT EXISTS idx_students_grade_section ON students(grade, section);
            CREATE INDEX IF NOT EXISTS idx_incidents_date ON incidents(incident_date);
            CREATE INDEX IF NOT EXISTS idx_incidents_student ON incidents(student_id);
            CREATE INDEX IF NOT EXISTS idx_incidents_offense ON incidents(offense_id);
            """
        )
        student_columns = {column["name"] for column in all_rows(conn, "PRAGMA table_info(students)")}
        if "password_hash" not in student_columns:
            conn.execute("ALTER TABLE students ADD COLUMN password_hash TEXT NOT NULL DEFAULT ''")
        sync_offense_catalog(conn)
    seed_if_empty()
    with db() as conn:
        ensure_student_passwords(conn)


def sync_offense_catalog(conn: sqlite3.Connection) -> None:
    for legacy_code, official_code in LEGACY_OFFENSE_CODE_MAP.items():
        conn.execute(
            "UPDATE offenses SET code = ? WHERE code = ?",
            (f"LEGACY-{legacy_code}", legacy_code),
        )

    for code, name, category, severity_points, recommended_action in OFFICIAL_OFFENSES:
        conn.execute(
            """
            INSERT INTO offenses (code, name, category, severity_points, recommended_action, active)
            VALUES (?, ?, ?, ?, ?, 1)
            ON CONFLICT(code) DO UPDATE SET
                name = excluded.name,
                category = excluded.category,
                severity_points = excluded.severity_points,
                recommended_action = excluded.recommended_action,
                active = 1,
                updated_at = CURRENT_TIMESTAMP
            """,
            (code, name, category, severity_points, recommended_action),
        )

    for legacy_code, official_code in LEGACY_OFFENSE_CODE_MAP.items():
        legacy = one(conn, "SELECT id FROM offenses WHERE code = ?", (f"LEGACY-{legacy_code}",))
        official = one(conn, "SELECT id FROM offenses WHERE code = ?", (official_code,))
        if not legacy or not official:
            continue
        conn.execute(
            "UPDATE incidents SET offense_id = ? WHERE offense_id = ?",
            (official["id"], legacy["id"]),
        )
        conn.execute("DELETE FROM offenses WHERE id = ?", (legacy["id"],))


def seed_if_empty() -> None:
    with db() as conn:
        if one(conn, "SELECT COUNT(*) AS count FROM students")["count"]:
            return

        first_names = [
            "Alyssa", "Marco", "Janelle", "Rafael", "Bianca", "Luis", "Trisha", "Enzo",
            "Mika", "Andre", "Sofia", "Nico", "Kath", "Paolo", "Lara", "Miguel",
            "Iya", "Jerome", "Camille", "Bryan", "Dana", "Anton", "Rhea", "Josh",
            "Sam", "Elaine", "Carlo", "Mara", "Francis", "Bea", "Kyle", "Nina",
            "Gio", "Clara", "Sean", "Ella", "Ken", "Jasmine", "Leo", "Mae",
            "Jan", "Mia", "Rico", "Angel", "Cedric", "Faith", "Ivan", "Alexa",
            "Noah", "Yna",
        ]
        last_names = [
            "Bautista", "Cruz", "Garcia", "Lim", "Reyes", "Santos", "Torres", "Villanueva",
            "Aquino", "Castillo", "Dela Cruz", "Flores", "Gonzales", "Lopez", "Mendoza", "Navarro",
            "Ortega", "Perez", "Ramos", "Sy", "Tan", "Valdez", "Yu", "Zamora",
            "Abad", "Bravo", "Chua", "Diaz", "Evangelista", "Fernandez", "Go", "Hernandez",
            "Ilagan", "Jacinto", "Kho", "Lazaro", "Morales", "Natividad", "Ocampo", "Pascual",
            "Quiambao", "Robles", "Salazar", "Tiu", "Uy", "Velasco", "Wong", "Xavier",
            "Yap", "Zulueta",
        ]
        advisers = {
            "7": "Ms. Reyes",
            "8": "Mr. Santos",
            "9": "Ms. Cruz",
            "10": "Mr. Flores",
            "11": "Ms. Lim",
            "12": "Mr. Garcia",
        }
        student_rows = []
        idx = 1
        for grade in ["7", "8", "9", "10", "11", "12"]:
            for section in ["A", "B", "C", "D"]:
                for _ in range(5):
                    if idx > 100:
                        break
                    first = first_names[(idx - 1) % len(first_names)]
                    last = last_names[(idx * 3 - 1) % len(last_names)]
                    student_rows.append(
                        (
                            f"RTS-{idx:05d}",
                            first,
                            last,
                            grade,
                            section,
                            advisers[grade],
                            "Active",
                        )
                    )
                    idx += 1
        conn.executemany(
            """
            INSERT INTO students (student_no, first_name, last_name, grade, section, adviser, status)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            student_rows,
        )

        student_ids = [row["id"] for row in all_rows(conn, "SELECT id FROM students ORDER BY id")]
        official_codes = [offense[0] for offense in OFFICIAL_OFFENSES]
        placeholders = ", ".join("?" for _ in official_codes)
        offense_ids = [
            row["id"]
            for row in all_rows(
                conn,
                f"SELECT id FROM offenses WHERE code IN ({placeholders}) ORDER BY code",
                tuple(official_codes),
            )
        ]
        statuses = ["Open", "Parent Notified", "Under Review", "Resolved"]
        reporters = ["Guidance Office", "Class Adviser", "Subject Teacher", "Prefect of Discipline"]
        start = date.today().replace(month=1, day=8)
        random.seed(20260723)
        incidents = []
        for i in range(160):
            incident_date = start + timedelta(days=random.randint(0, 195))
            offense_id = offense_ids[i % len(offense_ids)]
            action = one(conn, "SELECT recommended_action FROM offenses WHERE id = ?", (offense_id,))["recommended_action"]
            incidents.append(
                (
                    f"INC-{i + 1:05d}",
                    incident_date.isoformat(),
                    student_ids[(i * 11 + random.randint(0, 12)) % len(student_ids)],
                    offense_id,
                    action,
                    statuses[(i * 3) % len(statuses)],
                    reporters[i % len(reporters)],
                    "Seed sample incident for testing.",
                )
            )
        conn.executemany(
            """
            INSERT INTO incidents (
                incident_no, incident_date, student_id, offense_id, action_taken, status, reported_by, notes
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            incidents,
        )


@app.on_event("startup")
def startup() -> None:
    init_db()


@app.get("/api/health")
def health() -> dict:
    return {"status": "ok", "name": APP_NAME}


@app.post("/api/student-portal/login")
def student_portal_login(payload: StudentLoginIn) -> dict:
    with db() as conn:
        student = one(
            conn,
            f"SELECT {STUDENT_COLUMNS}, password_hash FROM students WHERE lower(student_no) = lower(?)",
            (payload.username.strip(),),
        )
        if not student or not verify_password(payload.password, student["password_hash"]):
            raise HTTPException(status_code=401, detail="Invalid Student ID or password")
        if student["status"] != "Active":
            raise HTTPException(status_code=403, detail="This student portal account is inactive")
        return {
            "access_token": encode_portal_token(student["id"]),
            "token_type": "bearer",
            "student": {key: value for key, value in student.items() if key != "password_hash"},
        }


@app.get("/api/student-portal/me")
def student_portal_me(student_id: int = Depends(portal_student_id)) -> dict:
    with db() as conn:
        student = one(
            conn,
            f"SELECT {STUDENT_COLUMNS} FROM students WHERE id = ?",
            (student_id,),
        )
        if not student or student["status"] != "Active":
            raise HTTPException(status_code=401, detail="This student portal account is unavailable")
        incidents = all_rows(
            conn,
            """
            SELECT i.id, i.incident_no, i.incident_date, i.action_taken, i.status,
                   i.reported_by, i.notes, o.code AS offense_code, o.name AS offense_name,
                   o.category, o.severity_points
            FROM incidents i
            JOIN offenses o ON o.id = i.offense_id
            WHERE i.student_id = ?
            ORDER BY i.incident_date DESC, i.id DESC
            """,
            (student_id,),
        )
        totals = one(
            conn,
            """
            SELECT COUNT(i.id) AS total_incidents,
                   COALESCE(SUM(CASE WHEN o.category = 'Minor' THEN 1 ELSE 0 END), 0) AS minor_incidents,
                   COALESCE(SUM(CASE WHEN o.category = 'Major' THEN 1 ELSE 0 END), 0) AS major_incidents,
                   COALESCE(SUM(o.severity_points), 0) AS severity_points
            FROM incidents i
            JOIN offenses o ON o.id = i.offense_id
            WHERE i.student_id = ?
            """,
            (student_id,),
        )
        return {"student": student, "totals": totals, "incidents": incidents}


@app.get("/api/students")
def list_students(
    search: str = "",
    grade: str = "",
    section: str = "",
    status: str = "",
    limit: int = Query(500, ge=1, le=1000),
) -> list[dict]:
    query = f"SELECT {STUDENT_COLUMNS} FROM students WHERE 1=1"
    params: list[str | int] = []
    if search:
        query += " AND (student_no LIKE ? OR first_name LIKE ? OR last_name LIKE ?)"
        like = f"%{search}%"
        params.extend([like, like, like])
    if grade:
        query += " AND grade = ?"
        params.append(grade)
    if section:
        query += " AND section = ?"
        params.append(section)
    if status:
        query += " AND status = ?"
        params.append(status)
    query += " ORDER BY CAST(grade AS INTEGER), section, last_name, first_name LIMIT ?"
    params.append(limit)
    with db() as conn:
        return all_rows(conn, query, tuple(params))


@app.post("/api/students", status_code=201)
def create_student(payload: StudentIn) -> dict:
    initial_password = default_student_password(payload.student_no, payload.last_name)
    with db() as conn:
        try:
            cur = conn.execute(
                """
                INSERT INTO students (
                    student_no, first_name, last_name, grade, section, adviser, status, password_hash
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    payload.student_no,
                    payload.first_name,
                    payload.last_name,
                    payload.grade,
                    payload.section,
                    payload.adviser,
                    payload.status,
                    hash_password(initial_password),
                ),
            )
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=409, detail="Student number already exists")
        student = one(
            conn,
            f"SELECT {STUDENT_COLUMNS} FROM students WHERE id = ?",
            (cur.lastrowid,),
        )
        return {**student, "username": payload.student_no, "initial_password": initial_password}


@app.post("/api/students/import")
async def import_students(file: UploadFile = File(...)) -> dict:
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an Excel .xlsx file")
    headers, records = parse_workbook(await file.read())
    require_import_headers(headers, {"student_no", "first_name", "last_name", "grade", "section"})

    imported = 0
    skipped = 0
    errors: list[dict] = []
    valid_statuses = {"active": "Active", "inactive": "Inactive"}
    seen: set[str] = set()

    with db() as conn:
        for row_number, row in records:
            student_no = clean_text(row.get("student_no"))
            first_name = clean_text(row.get("first_name"))
            last_name = clean_text(row.get("last_name"))
            grade = clean_text(row.get("grade"))
            section = clean_text(row.get("section"))
            adviser = clean_text(row.get("adviser"))
            status_text = clean_text(row.get("status") or "Active")
            status = valid_statuses.get(status_text.lower())

            missing = [
                label
                for label, value in (
                    ("Student No.", student_no),
                    ("First Name", first_name),
                    ("Last Name", last_name),
                    ("Grade", grade),
                    ("Section", section),
                )
                if not value
            ]
            if missing:
                errors.append({"row": row_number, "message": f"Missing {', '.join(missing)}"})
                continue
            if not status:
                errors.append({"row": row_number, "message": "Status must be Active or Inactive"})
                continue

            duplicate_key = student_no.lower()
            if duplicate_key in seen or one(
                conn,
                "SELECT id FROM students WHERE lower(student_no) = lower(?)",
                (student_no,),
            ):
                skipped += 1
                continue

            conn.execute(
                """
                INSERT INTO students (
                    student_no, first_name, last_name, grade, section, adviser, status, password_hash
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    student_no,
                    first_name,
                    last_name,
                    grade,
                    section,
                    adviser,
                    status,
                    hash_password(default_student_password(student_no, last_name)),
                ),
            )
            seen.add(duplicate_key)
            imported += 1

    return {
        "imported": imported,
        "skipped": skipped,
        "error_count": len(errors),
        "errors": errors[:10],
        "total_rows": len(records),
    }


@app.put("/api/students/{student_id}")
def update_student(student_id: int, payload: StudentIn) -> dict:
    initial_password = default_student_password(payload.student_no, payload.last_name)
    with db() as conn:
        if not one(conn, "SELECT id FROM students WHERE id = ?", (student_id,)):
            raise HTTPException(status_code=404, detail="Student not found")
        conn.execute(
            """
            UPDATE students
            SET student_no = ?, first_name = ?, last_name = ?, grade = ?, section = ?,
                adviser = ?, status = ?, password_hash = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                payload.student_no,
                payload.first_name,
                payload.last_name,
                payload.grade,
                payload.section,
                payload.adviser,
                payload.status,
                hash_password(initial_password),
                student_id,
            ),
        )
        student = one(
            conn,
            f"SELECT {STUDENT_COLUMNS} FROM students WHERE id = ?",
            (student_id,),
        )
        return {**student, "username": payload.student_no, "initial_password": initial_password}


@app.post("/api/students/bulk-delete")
def bulk_delete_students(payload: BulkDeleteIn) -> dict:
    student_ids = list(dict.fromkeys(payload.ids))
    placeholders = ", ".join("?" for _ in student_ids)

    with db() as conn:
        selected = all_rows(
            conn,
            f"SELECT id FROM students WHERE id IN ({placeholders})",
            tuple(student_ids),
        )
        if len(selected) != len(student_ids):
            raise HTTPException(
                status_code=404,
                detail="One or more selected students were not found. Refresh the list and try again",
            )

        protected = all_rows(
            conn,
            f"""
            SELECT s.student_no, s.first_name, s.last_name, COUNT(i.id) AS incident_count
            FROM students s
            JOIN incidents i ON i.student_id = s.id
            WHERE s.id IN ({placeholders})
            GROUP BY s.id
            ORDER BY s.last_name, s.first_name
            """,
            tuple(student_ids),
        )
        if protected:
            preview = ", ".join(
                f"{row['first_name']} {row['last_name']} ({row['student_no']})"
                for row in protected[:3]
            )
            remaining = len(protected) - 3
            if remaining > 0:
                preview += f", and {remaining} more"
            raise HTTPException(
                status_code=409,
                detail=(
                    f"{len(protected)} selected student{' has' if len(protected) == 1 else 's have'} "
                    f"incident records and cannot be deleted: {preview}. No students were deleted"
                ),
            )

        conn.execute(
            f"DELETE FROM students WHERE id IN ({placeholders})",
            tuple(student_ids),
        )
        return {"deleted": len(student_ids)}


@app.delete("/api/students/{student_id}", status_code=204)
def delete_student(student_id: int) -> None:
    with db() as conn:
        try:
            conn.execute("DELETE FROM students WHERE id = ?", (student_id,))
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=409, detail="Student has incidents and cannot be deleted")


@app.get("/api/offenses")
def list_offenses(category: str = "", active_only: bool = False) -> list[dict]:
    query = "SELECT * FROM offenses WHERE 1=1"
    params: list[str | int] = []
    if category:
        query += " AND category = ?"
        params.append(category)
    if active_only:
        query += " AND active = 1"
    query += " ORDER BY category DESC, code"
    with db() as conn:
        return all_rows(conn, query, tuple(params))


@app.post("/api/offenses", status_code=201)
def create_offense(payload: OffenseIn) -> dict:
    with db() as conn:
        try:
            cur = conn.execute(
                """
                INSERT INTO offenses (code, name, category, severity_points, recommended_action, active)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    payload.code,
                    payload.name,
                    payload.category,
                    payload.severity_points,
                    payload.recommended_action,
                    int(payload.active),
                ),
            )
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=409, detail="Offense code already exists")
        return one(conn, "SELECT * FROM offenses WHERE id = ?", (cur.lastrowid,))


@app.put("/api/offenses/{offense_id}")
def update_offense(offense_id: int, payload: OffenseIn) -> dict:
    with db() as conn:
        if not one(conn, "SELECT id FROM offenses WHERE id = ?", (offense_id,)):
            raise HTTPException(status_code=404, detail="Offense not found")
        conn.execute(
            """
            UPDATE offenses
            SET code = ?, name = ?, category = ?, severity_points = ?, recommended_action = ?,
                active = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                payload.code,
                payload.name,
                payload.category,
                payload.severity_points,
                payload.recommended_action,
                int(payload.active),
                offense_id,
            ),
        )
        return one(conn, "SELECT * FROM offenses WHERE id = ?", (offense_id,))


@app.post("/api/offenses/bulk-delete")
def bulk_delete_offenses(payload: BulkDeleteIn) -> dict:
    offense_ids = list(dict.fromkeys(payload.ids))
    placeholders = ", ".join("?" for _ in offense_ids)

    with db() as conn:
        selected = all_rows(
            conn,
            f"SELECT id FROM offenses WHERE id IN ({placeholders})",
            tuple(offense_ids),
        )
        if len(selected) != len(offense_ids):
            raise HTTPException(
                status_code=404,
                detail="One or more selected offenses were not found. Refresh the list and try again",
            )

        protected = all_rows(
            conn,
            f"""
            SELECT o.code, o.name, COUNT(i.id) AS incident_count
            FROM offenses o
            JOIN incidents i ON i.offense_id = o.id
            WHERE o.id IN ({placeholders})
            GROUP BY o.id
            ORDER BY o.code
            """,
            tuple(offense_ids),
        )
        if protected:
            preview = ", ".join(
                f"{row['code']} - {row['name']}"
                for row in protected[:3]
            )
            remaining = len(protected) - 3
            if remaining > 0:
                preview += f", and {remaining} more"
            raise HTTPException(
                status_code=409,
                detail=(
                    f"{len(protected)} selected offense{' has' if len(protected) == 1 else 's have'} "
                    f"incident records and cannot be deleted: {preview}. No offenses were deleted"
                ),
            )

        conn.execute(
            f"DELETE FROM offenses WHERE id IN ({placeholders})",
            tuple(offense_ids),
        )
        return {"deleted": len(offense_ids)}


@app.delete("/api/offenses/{offense_id}", status_code=204)
def delete_offense(offense_id: int) -> None:
    with db() as conn:
        try:
            conn.execute("DELETE FROM offenses WHERE id = ?", (offense_id,))
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=409, detail="Offense has incidents and cannot be deleted")


def next_incident_no(conn: sqlite3.Connection) -> str:
    row = one(
        conn,
        """
        SELECT COALESCE(MAX(CAST(SUBSTR(incident_no, 5) AS INTEGER)), 0) + 1 AS next_no
        FROM incidents
        WHERE incident_no LIKE 'INC-%'
        """,
    )
    return f"INC-{row['next_no']:05d}"


@app.get("/api/incidents")
def list_incidents(
    search: str = "",
    month: str = "",
    category: str = "",
    status: str = "",
    limit: int = Query(500, ge=1, le=1000),
) -> list[dict]:
    query = """
        SELECT i.*, s.student_no, s.first_name, s.last_name, s.grade, s.section,
               o.code AS offense_code, o.name AS offense_name, o.category, o.severity_points
        FROM incidents i
        JOIN students s ON s.id = i.student_id
        JOIN offenses o ON o.id = i.offense_id
        WHERE 1=1
    """
    params: list[str | int] = []
    if search:
        like = f"%{search}%"
        query += " AND (i.incident_no LIKE ? OR s.student_no LIKE ? OR s.first_name LIKE ? OR s.last_name LIKE ? OR o.name LIKE ?)"
        params.extend([like, like, like, like, like])
    if month:
        query += " AND strftime('%Y-%m', i.incident_date) = ?"
        params.append(month)
    if category:
        query += " AND o.category = ?"
        params.append(category)
    if status:
        query += " AND i.status = ?"
        params.append(status)
    query += " ORDER BY i.incident_date DESC, i.id DESC LIMIT ?"
    params.append(limit)
    with db() as conn:
        return all_rows(conn, query, tuple(params))


@app.post("/api/incidents", status_code=201)
def create_incident(payload: IncidentIn) -> dict:
    with db() as conn:
        if not one(conn, "SELECT id FROM students WHERE id = ?", (payload.student_id,)):
            raise HTTPException(status_code=404, detail="Student not found")
        offense = one(conn, "SELECT * FROM offenses WHERE id = ?", (payload.offense_id,))
        if not offense:
            raise HTTPException(status_code=404, detail="Offense not found")
        incident_date = (payload.incident_date or date.today()).isoformat()
        action_taken = payload.action_taken or offense["recommended_action"]
        cur = conn.execute(
            """
            INSERT INTO incidents (
                incident_no, incident_date, student_id, offense_id, action_taken, status, reported_by, notes
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                next_incident_no(conn),
                incident_date,
                payload.student_id,
                payload.offense_id,
                action_taken,
                payload.status,
                payload.reported_by,
                payload.notes,
            ),
        )
        return one(conn, "SELECT * FROM incidents WHERE id = ?", (cur.lastrowid,))


@app.post("/api/incidents/import")
async def import_incidents(file: UploadFile = File(...)) -> dict:
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an Excel .xlsx file")
    headers, records = parse_workbook(await file.read())
    require_import_headers(headers, {"incident_date", "student_no", "offense_code"})

    imported = 0
    errors: list[dict] = []
    valid_statuses = {
        "open": "Open",
        "parent notified": "Parent Notified",
        "under review": "Under Review",
        "resolved": "Resolved",
    }

    with db() as conn:
        for row_number, row in records:
            student_no = clean_text(row.get("student_no"))
            offense_code = clean_text(row.get("offense_code"))
            status_text = clean_text(row.get("status") or "Open")
            status = valid_statuses.get(status_text.lower())

            student = one(
                conn,
                "SELECT id FROM students WHERE lower(student_no) = lower(?)",
                (student_no,),
            )
            offense = one(
                conn,
                "SELECT id, recommended_action FROM offenses WHERE lower(code) = lower(?)",
                (offense_code,),
            )
            if not student_no:
                errors.append({"row": row_number, "message": "Student No. is required"})
                continue
            if not offense_code:
                errors.append({"row": row_number, "message": "Offense Code is required"})
                continue
            if not student:
                errors.append({"row": row_number, "message": f"Student {student_no} was not found"})
                continue
            if not offense:
                errors.append({"row": row_number, "message": f"Offense code {offense_code} was not found"})
                continue
            if not status:
                errors.append({
                    "row": row_number,
                    "message": "Status must be Open, Parent Notified, Under Review, or Resolved",
                })
                continue
            try:
                incident_date = import_date(row.get("incident_date")).isoformat()
            except (TypeError, ValueError) as exc:
                errors.append({"row": row_number, "message": str(exc)})
                continue

            action_taken = clean_text(row.get("action_taken")) or offense["recommended_action"]
            conn.execute(
                """
                INSERT INTO incidents (
                    incident_no, incident_date, student_id, offense_id,
                    action_taken, status, reported_by, notes
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    next_incident_no(conn),
                    incident_date,
                    student["id"],
                    offense["id"],
                    action_taken,
                    status,
                    clean_text(row.get("reported_by")),
                    clean_text(row.get("notes")),
                ),
            )
            imported += 1

    return {
        "imported": imported,
        "skipped": 0,
        "error_count": len(errors),
        "errors": errors[:10],
        "total_rows": len(records),
    }


@app.put("/api/incidents/{incident_id}")
def update_incident(incident_id: int, payload: IncidentIn) -> dict:
    with db() as conn:
        if not one(conn, "SELECT id FROM incidents WHERE id = ?", (incident_id,)):
            raise HTTPException(status_code=404, detail="Incident not found")
        if not one(conn, "SELECT id FROM students WHERE id = ?", (payload.student_id,)):
            raise HTTPException(status_code=404, detail="Student not found")
        offense = one(conn, "SELECT * FROM offenses WHERE id = ?", (payload.offense_id,))
        if not offense:
            raise HTTPException(status_code=404, detail="Offense not found")
        conn.execute(
            """
            UPDATE incidents
            SET incident_date = ?, student_id = ?, offense_id = ?, action_taken = ?,
                status = ?, reported_by = ?, notes = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                (payload.incident_date or date.today()).isoformat(),
                payload.student_id,
                payload.offense_id,
                payload.action_taken or offense["recommended_action"],
                payload.status,
                payload.reported_by,
                payload.notes,
                incident_id,
            ),
        )
        return one(conn, "SELECT * FROM incidents WHERE id = ?", (incident_id,))


@app.post("/api/incidents/bulk-delete")
def bulk_delete_incidents(payload: BulkDeleteIn) -> dict:
    incident_ids = list(dict.fromkeys(payload.ids))
    placeholders = ", ".join("?" for _ in incident_ids)

    with db() as conn:
        selected = all_rows(
            conn,
            f"SELECT id FROM incidents WHERE id IN ({placeholders})",
            tuple(incident_ids),
        )
        if len(selected) != len(incident_ids):
            raise HTTPException(
                status_code=404,
                detail="One or more selected incidents were not found. Refresh the list and try again",
            )
        conn.execute(
            f"DELETE FROM incidents WHERE id IN ({placeholders})",
            tuple(incident_ids),
        )
        return {"deleted": len(incident_ids)}


@app.delete("/api/incidents/{incident_id}", status_code=204)
def delete_incident(incident_id: int) -> None:
    with db() as conn:
        conn.execute("DELETE FROM incidents WHERE id = ?", (incident_id,))


def report_year(value: int | None) -> int:
    return value or date.today().year


@app.get("/api/reports/dashboard")
def dashboard_report(year: int | None = None) -> dict:
    year = report_year(year)
    with db() as conn:
        totals = one(
            conn,
            """
            SELECT
                COUNT(*) AS total_incidents,
                SUM(CASE WHEN o.category = 'Minor' THEN 1 ELSE 0 END) AS minor_incidents,
                SUM(CASE WHEN o.category = 'Major' THEN 1 ELSE 0 END) AS major_incidents,
                COALESCE(SUM(o.severity_points), 0) AS severity_points,
                SUM(CASE WHEN i.status = 'Open' THEN 1 ELSE 0 END) AS open_cases,
                COUNT(DISTINCT i.student_id) AS students_with_incidents
            FROM incidents i
            JOIN offenses o ON o.id = i.offense_id
            WHERE strftime('%Y', i.incident_date) = ?
            """,
            (str(year),),
        )
        active_students = one(conn, "SELECT COUNT(*) AS count FROM students WHERE status = 'Active'")["count"]
        high_risk = one(
            conn,
            """
            SELECT COUNT(*) AS count
            FROM (
                SELECT i.student_id, SUM(o.severity_points) AS points
                FROM incidents i
                JOIN offenses o ON o.id = i.offense_id
                WHERE strftime('%Y', i.incident_date) = ?
                GROUP BY i.student_id
                HAVING points >= 12
            )
            """,
            (str(year),),
        )["count"]
        monthly = all_rows(
            conn,
            """
            SELECT strftime('%Y-%m', i.incident_date) AS month,
                   COUNT(*) AS total,
                   SUM(CASE WHEN o.category = 'Minor' THEN 1 ELSE 0 END) AS minor,
                   SUM(CASE WHEN o.category = 'Major' THEN 1 ELSE 0 END) AS major,
                   COALESCE(SUM(o.severity_points), 0) AS points
            FROM incidents i
            JOIN offenses o ON o.id = i.offense_id
            WHERE strftime('%Y', i.incident_date) = ?
            GROUP BY month
            ORDER BY month
            """,
            (str(year),),
        )
        top_students = all_rows(
            conn,
            """
            SELECT s.id, s.student_no, s.first_name, s.last_name, s.grade, s.section,
                   COUNT(i.id) AS incident_count, SUM(o.severity_points) AS points
            FROM incidents i
            JOIN students s ON s.id = i.student_id
            JOIN offenses o ON o.id = i.offense_id
            WHERE strftime('%Y', i.incident_date) = ?
            GROUP BY s.id
            ORDER BY points DESC, incident_count DESC
            LIMIT 10
            """,
            (str(year),),
        )
        top_offenses = all_rows(
            conn,
            """
            SELECT o.code, o.name, o.category, COUNT(i.id) AS incident_count
            FROM incidents i
            JOIN offenses o ON o.id = i.offense_id
            WHERE strftime('%Y', i.incident_date) = ?
            GROUP BY o.id
            ORDER BY incident_count DESC, o.name
            LIMIT 10
            """,
            (str(year),),
        )
    return {
        "year": year,
        "totals": {**totals, "active_students": active_students, "high_risk_students": high_risk},
        "monthly": monthly,
        "top_students": top_students,
        "top_offenses": top_offenses,
    }


@app.get("/api/reports/student-summary")
def student_summary(year: int | None = None) -> list[dict]:
    year = report_year(year)
    with db() as conn:
        return all_rows(
            conn,
            """
            SELECT s.id, s.student_no, s.first_name, s.last_name, s.grade, s.section,
                   COUNT(i.id) AS total_offenses,
                   SUM(CASE WHEN o.category = 'Minor' THEN 1 ELSE 0 END) AS minor_offenses,
                   SUM(CASE WHEN o.category = 'Major' THEN 1 ELSE 0 END) AS major_offenses,
                   COALESCE(SUM(o.severity_points), 0) AS severity_points,
                   MAX(i.incident_date) AS last_incident_date,
                   CASE
                       WHEN COALESCE(SUM(o.severity_points), 0) >= 12 THEN 'High'
                       WHEN COALESCE(SUM(o.severity_points), 0) >= 6 THEN 'Monitor'
                       WHEN COUNT(i.id) > 0 THEN 'Low'
                       ELSE 'None'
                   END AS risk_band
            FROM students s
            LEFT JOIN incidents i ON i.student_id = s.id AND strftime('%Y', i.incident_date) = ?
            LEFT JOIN offenses o ON o.id = i.offense_id
            GROUP BY s.id
            ORDER BY severity_points DESC, total_offenses DESC, s.grade, s.section, s.last_name
            """,
            (str(year),),
        )
